#!/usr/bin/env python
# -*- coding: utf-8 -*-

import urllib
from bs4 import BeautifulSoup
import requests
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
import json
import time
import os
import openpyxl

data_crawl = {}

# class for company object
class item:
    def __init__(self, name, tax_code, address, legal_representative, issue_date, date_of_operation, phone, url):
        self.name = name
        self.tax_code = tax_code
        self.address = address
        self.legal_representative = legal_representative
        self.issue_date = issue_date
        self.date_of_operation = date_of_operation
        self.phone = phone
        self.url = url
# DONE


# get all link company
def get_all_url_companys(url_base, start, end):

    links = []

    for i in range(start, end + 1):
        # print "Page number: ", i
        url = url_base % (urllib.quote(str(i)))
        r = requests.get(url)
        data = r.text
        soup = BeautifulSoup(data, 'html.parser')
        for company in soup.find_all('div', attrs={'class': "search-results"}):
            link = company.find('a')
            links.append(link.get('href'))
    return links
# DONE


# process data for every company
def get_infor_per_company(url):
    r = requests.get(url)
    data = r.text
    soup = BeautifulSoup(data, 'html.parser')
    info = soup.find('div', attrs={'class': "jumbotron"})
    name = ''
    try:
        name = info.find('span').text
    except:
        pass

    tax_code = ''
    phone = ''

    img_src = info.findAll('img')
    if (len(img_src) == 2):
        tax_code = img_src[0].get('src')
        phone = img_src[1].get('src')
    elif (len(img_src) == 1):
        tax_code = img_src[0].get('src')


    company_info = info.text

    company_modifies = ""
    found = False
    multiSpace = False

    for i in company_info:
        if ord(i) == 32:
            if (found == True):
                if (multiSpace == False):
                    company_modifies += ':'
                    multiSpace = True
            else:
                multiSpace = False
            if found == False:
                found = True
                company_modifies += i

        elif ord(i) != 13 and ord(i) != 10:
            found = False
            company_modifies += i


    company_modifies = company_modifies.split(':')

    address = ''
    legal_representative = ''
    issue_date = ''
    date_of_operation = ''

    for i in range(0, len(company_modifies) - 1):
        length = len(company_modifies[i])
        value = company_modifies[i + 1][1:]
        if length == 7:  # Dia chi and ord(company_modifies[i][0] == 272)
            address = value
        elif length == 18:
            if company_modifies[i][0] == 'N':
                issue_date = value
            else:
                legal_representative = value
        elif length == 14:
            date_of_operation = value

    # print name
    # print address
    # print legal_representative
    # print issue_date
    # print date_of_operation
    # print tax_code
    # print phone

    ret = item(name, tax_code, address, legal_representative, issue_date, date_of_operation, phone, url)
    data_crawl['item'].append(ret.__dict__)
    return ret
# DONE


# in ra tat ca cac url cua cong ty file out_txt// tu page start -> end
def print_link_txt(url_base, start_page, end_page, out_txt):

    links = get_all_url_companys(url_base, start_page, end_page)
    print_data_txt(links, out_txt)

    return 0
# DONE


# doc url cua cac cong ty trong file inp_txt roi crawl lay thong tin sau do in ra file out_json
def print_info_company_json(inp_txt, out_json):

    urls = read_url_company_from_txt(inp_txt)
    f = open(out_json, "w")

    data_crawl['item'] = []
    for i in range(0, len(urls)):
        print i, " ", urls[i]
        get_infor_per_company(urls[i])

    json.dump(data_crawl, f, ensure_ascii=False)
    f.close()

    return 0
# DONE


# write data (type list) into file .txt
def print_data_txt(data, out_txt):
    f = open(out_txt, "w")
    for i in data:
        f.write(i + "\n")
    f.close()
# DONE


# write data (type dict) into file .json
def print_data_json(data, out_json):
    f = open(out_json, "w")
    json.dump(data, f, ensure_ascii=False)
    f.close()
# DONE


# doc url cua cac cong ty trong file txt
def read_url_company_from_txt(inp_txt):
    urls = []
    f = open(inp_txt, "r")
    for url in f:
        urls.append(url[:-1])
    return urls
# DONE


# doc partern cua cac url tren tung don vi trong file xlsx
def read_url_company_from_xlsx(inp_xlsx, col_index_url, col_index_page, num_row):
    url_bases = []
    num_page =  []

    book = openpyxl.load_workbook(inp_xlsx)
    sheet = book.active
    for row in range(2, num_row + 1):  # ignore 1st row which is preserved for column title
        data_url = sheet.cell(row=row, column=col_index_url).value
        url_bases.append(data_url)
        data_page = sheet.cell(row=row, column=col_index_page).value
        num_page.append(data_page)

    return url_bases, num_page
