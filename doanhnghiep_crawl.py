#!/usr/bin/env python
# -*- coding: utf-8 -*-

import urllib
from bs4 import BeautifulSoup
import requests
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
import json
import time
import os
import openpyxl
data_crawl = {}

data_crawl['item'] = []

# class for company object
class item:
    def __init__(self, tax_code, tc_iss_date, tc_exp_date, name_in_vnese, name_in_tran,
                 tax_reg_ad, tel_num, head_off_add, tax_rec_add, reg_cert_date, reg_cert_num, reg_cert_ad, fin_year,
                 tax_reg_rec_date, date_com_op, chart_cap, num_employee, acc_model, vat_cal_method, own_name, own_add,
                 dir_name, dir_add, chief_acc, chief_acc_add, main_busi_line, paid_tax, url):

        self.tax_code = tax_code                    # tax_code # Mã số ĐTNT
        self.tc_iss_date = tc_iss_date              # tax_code_issue_date # Ngày cấp
        self.tc_exp_date = tc_exp_date              # tax_code_expired_date # Ngày đóng MST
        self.name_in_vnese = name_in_vnese          # name_in_vnese # Tên chính thức
        self.name_in_tran = name_in_tran            # name_in_transaction # Tên giao dịch
        self.tax_reg_ad = tax_reg_ad                # tax_registration_administrative # Nơi đăng ký quản lý
        self.tel_num = tel_num                      # telephone_number # Điện thoại
        self.head_off_add = head_off_add            # head_office_address # Địa chỉ trụ sở
        self.tax_rec_add = tax_rec_add              # tax_received_address # Địa chỉ nhận thông báo thuế
        self.reg_cert_date = reg_cert_date          # enterprise_registration_certificate_date # QĐTL/Ngày cấp
        self.reg_cert_num = reg_cert_num            # enterprise_registration_certificate_number # GPKD/Ngày cấp
        self.reg_cert_ad = reg_cert_ad              # enterprise_registration_certificate_administrative # Cơ quan cấp
        self.fin_year = fin_year                    # finalcial_year # Năm tài chính
        self.tax_reg_rec_date = tax_reg_rec_date    # tax_registration_form_received_date # Ngày nhận TK
        self.date_com_op = date_com_op              # date_of_commencement_of_operation # Ngày bắt đầu HĐ
        self.chart_cap = chart_cap                  # charter_capital # Vốn điều lệ
        self.num_employee = num_employee            # total_number_of_employee# Tổng số lao động
        self.acc_model = acc_model                  # accounting_model # Hình thức h.toán
        self.vat_cal_method = vat_cal_method        # vat_calculation_method # PP tính thuế GTGT
        self.own_name = own_name                    # owner_name # Chủ sở hữu : Đỗ Văn Bắc
        self.own_add = own_add                      # owner_address # Địa chỉ chủ sở hữu
        self.dir_name = dir_name                    # director_name # Tên giám đốc
        self.dir_add = dir_add                      # director_address # Địa chỉ
        self.chief_acc = chief_acc                  # chief_accountant # Kế toán trưởng
        self.chief_acc_add = chief_acc_add          # chief_accountant_address # Địa chỉ
        self.main_busi_line = main_busi_line        # main_business_line # Ngành nghề chính
        self.paid_tax = paid_tax                    # paid_taxes  # Loại thuế phải nộp
        self.url = url                              # url # link chua thong tin cong ty

        # skip # Nơi đăng ký nộp thuế
        # skip #  Điện thoại / Fax
        # skip # C.Q ra quyết định
        # skip # Mã số hiện thời
        # skip # Cấp Chương loại khoản
# DONE


# get all link company
def get_all_url_companys(url_base, start, end):
    print "url_base: ", url_base
    links = []
    for i in range(start, end + 1):
        len_page = 0
        while len_page < 1000:
            print "Page number: ", i
            url = url_base % (urllib.quote(str(i)))
            r = requests.get(url)
            data = r.text
            soup = BeautifulSoup(data, 'html.parser')
            # print "Len data: ", len(data)
            len_page = len(data)
            for company in soup.find_all('div', attrs={'class': "news-v3 bg-color-white"}):
                link = company.find('a')
                url = 'https://thongtindoanhnghiep.co' + link.get('href')
                links.append(url)
                # print "Link: ", url

    return links
# DONE


# process data for every company
def get_infor_per_company(url):
    len_page = 0
    while len_page < 1000:
        r = requests.get(url)

        data = r.text
        len_page = len(data)

        if len_page >= 1000:
            # print "Url: ", url, " len page: ", len_page
            soup = BeautifulSoup(data, 'html.parser')

            info = soup.find('table', attrs={'class': "table table-striped table-bordered table-responsive table-details"})
            tr = info.findAll('tr')
            # print "Len tr:", len(tr)
            com_info = []
            for i in tr:
                td = i.findAll('td')
                th = i.findAll('th')
                for i in range(0, len(td)):
                    # print th[i].text, ":", td[i].text, " ", len(td[i].text)
                    com_info.append(td[i].text)

            # print len(com_info)
            # for i in range(0, len(com_info)):
            #     info = com_info[i]
            #     print i, " ", info

            ret = item(tax_code=com_info[0], tc_iss_date=com_info[1], tc_exp_date=com_info[2], name_in_vnese=com_info[3], name_in_tran=com_info[4],
                       tax_reg_ad=com_info[5], tel_num=com_info[6], head_off_add=com_info[7], tax_rec_add=com_info[10], reg_cert_date=com_info[11],
                       reg_cert_num=com_info[13], reg_cert_ad=com_info[14], fin_year=com_info[15], tax_reg_rec_date=com_info[17], date_com_op=com_info[18],
                       chart_cap=com_info[19], num_employee=com_info[20], acc_model=com_info[22], vat_cal_method=com_info[23], own_name=com_info[24],
                       own_add=com_info[25], dir_name=com_info[26], dir_add=com_info[27], chief_acc=com_info[28], chief_acc_add=com_info[29],
                       main_busi_line=com_info[30], paid_tax=com_info[31], url=url)

            data_crawl['item'].append(ret.__dict__)

# DONE


# in ra tat ca cac url cua cong ty file out_txt// tu page start -> end
def print_link_txt(url_base, start_page, end_page, out_txt):
    links = get_all_url_companys(url_base, start_page, end_page)
    print_data_txt(links, out_txt)
    return 0
# DONE


# doc url cua cac cong ty trong file inp_txt roi crawl lay thong tin sau do in ra file out_json
def print_info_company_json(inp_txt, out_json):

    urls = read_url_company_from_txt(inp_txt)
    f = open(out_json, "w")

    # data_crawl['item'] = []
    for i in range(0, len(urls)):
        # print i, " ", urls[i]
        get_infor_per_company(urls[i])

    json.dump(data_crawl, f, ensure_ascii=False)
    f.close()

    return 0
# DONE


# write data (type list) into file .txt
def print_data_txt(data, out_txt):
    f = open(out_txt, "w")
    for i in data:
        f.write(i + "\n")
    f.close()
# DONE


# write data (type dict) into file .json
def print_data_json(data, out_json):
    f = open(out_json, "w")
    json.dump(data, f, ensure_ascii=False)
    f.close()
# DONE


# doc url cua cac cong ty trong file txt
def read_url_company_from_txt(inp_txt):
    urls = []
    f = open(inp_txt, "r")
    for url in f:
        urls.append(url[:-1])
    return urls
# DONE


# doc partern cua cac url tren tung don vi trong file xlsx
# read_url_company_from_xlsx(ttdn.xlsx,2, 3, 563)
def read_url_company_from_xlsx(inp_xlsx, col_index_url, col_index_page, num_row):
    url_bases = []
    num_page = []

    book = openpyxl.load_workbook(inp_xlsx)
    sheet = book.active
    for row in range(2, num_row + 1):  # ignore 1st row which is preserved for column title
        data_url = sheet.cell(row=row, column=col_index_url).value
        url_bases.append(data_url)
        data_page = sheet.cell(row=row, column=col_index_page).value
        num_page.append(data_page)

    return url_bases, num_page

# chuyen doi format so sang 3 chu so: 1 --> 001; 10 -> 010 %3d
def convert_format_int(value):
    ret = ''
    if value < 10:
        ret += '00' + str(value)
    elif value < 100:
        ret += '0' + str(value)
    else:
        ret += str(value)
    return ret